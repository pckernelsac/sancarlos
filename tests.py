"""
Tests del Sistema Académico San Carlos
Ejecutar: python -m pytest tests.py -v
"""
import os
import re
# Forzar BD en memoria ANTES de importar la app, para que SQLAlchemy
# nunca toque el archivo real sancarlos.db
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
# Tests: sin rate limit en login (evita flakiness)
os.environ["LOGIN_RATE_LIMIT_MAX_FAILURES"] = "0"

import pytest
from starlette.testclient import TestClient

from app import create_app
from app.database import db
from app.models.user import User, RoleEnum, TeacherCourse
from app.models.student import Student
from app.models.academic import (
    Course, Term, Grade, Attendance, Behavior, EDA, EdaGrade,
    RegistroSemana, RegistroExamen, ParentResponsibility,
)
from app.services.grade_service import numeric_to_qualitative, upsert_grade
from app.services.excel_import_teachers import (
    import_teachers_from_excel,
    generate_teachers_template_excel,
    _cell_to_str,
    _parse_course_ids,
    _grados_validos,
)


# ─── Fixtures ────────────────────────────────────────────────────────

@pytest.fixture
def app():
    """Crea app con BD en memoria para tests."""
    application = create_app()
    db.create_all()
    yield application
    db.remove_session()
    db.drop_all()


@pytest.fixture
def client(app):
    return TestClient(app)


@pytest.fixture
def seed(app):
    """Crea datos básicos y retorna dict con IDs."""
    admin = User(username="admin", full_name="Administrador", role=RoleEnum.ADMIN)
    admin.set_password("admin1234")

    docente = User(username="docente", full_name="Prof. García", role=RoleEnum.DOCENTE)
    docente.set_password("docente1234")

    auxiliar = User(username="auxiliar", full_name="Aux. López", role=RoleEnum.AUXILIAR)
    auxiliar.set_password("auxiliar1234")

    db.session.add_all([admin, docente, auxiliar])
    db.session.flush()

    term = Term(nombre="I Bimestre", orden=1, anio=2026)
    db.session.add(term)
    db.session.flush()

    eda = EDA(term_id=term.id, nombre="EDA 1", orden=1)
    db.session.add(eda)
    db.session.flush()

    course = Course(nombre="Matemática", area="Matemática", nivel="PRIMARIA")
    db.session.add(course)
    db.session.flush()

    tc = TeacherCourse(user_id=docente.id, course_id=course.id)
    db.session.add(tc)

    student = Student(
        codigo="2026-001", nombres="Juan",
        apellido_paterno="PEREZ", apellido_materno="GOMEZ",
        nivel="PRIMARIA", grado="3", seccion="A", estado="ACTIVO",
    )
    db.session.add(student)
    db.session.commit()

    return {
        "admin_id": admin.id, "docente_id": docente.id, "auxiliar_id": auxiliar.id,
        "term_id": term.id, "course_id": course.id, "student_id": student.id,
        "eda_id": eda.id,
    }


def login(client, username, password):
    """Inicia sesión; TestClient mantiene cookies automáticamente."""
    page = client.get("/auth/login")
    m = re.search(r'name="csrf_token"\s+value="([^"]+)"', page.text)
    token = m.group(1) if m else ""
    return client.post(
        "/auth/login",
        data={"username": username, "password": password, "csrf_token": token},
        headers={"X-CSRFToken": token},
        follow_redirects=True,
    )


# ─── 1. Tests de creación de la app ─────────────────────────────────

class TestAppFactory:
    def test_app_creates(self, app):
        assert app is not None

    def test_routers_registered(self, app):
        """Verifica que las rutas principales existan."""
        paths = [r.path for r in app.routes if hasattr(r, "path")]
        # Verificar que al menos algunas rutas clave están registradas
        assert any("/auth" in p for p in paths)
        assert any("/students" in p for p in paths)
        assert any("/grades" in p for p in paths)


# ─── 2. Tests de modelos ────────────────────────────────────────────

class TestUserModel:
    def test_create_user(self, app, seed):
        user = db.session.get(User, seed["admin_id"])
        assert user.username == "admin"
        assert user.role == RoleEnum.ADMIN

    def test_password_hashing(self, app, seed):
        user = db.session.get(User, seed["admin_id"])
        assert user.check_password("admin1234") is True
        assert user.check_password("wrong") is False

    def test_has_role(self, app, seed):
        admin = db.session.get(User, seed["admin_id"])
        docente = db.session.get(User, seed["docente_id"])
        auxiliar = db.session.get(User, seed["auxiliar_id"])
        assert admin.has_role("ADMIN") is True
        assert docente.has_role("DOCENTE") is True
        assert auxiliar.has_role("AUXILIAR") is True
        assert docente.has_role("ADMIN") is False

    def test_can_grade_course_admin(self, app, seed):
        admin = db.session.get(User, seed["admin_id"])
        assert admin.can_grade_course(seed["course_id"]) is True

    def test_can_grade_course_docente(self, app, seed):
        docente = db.session.get(User, seed["docente_id"])
        assert docente.can_grade_course(seed["course_id"]) is True
        assert docente.can_grade_course(99999) is False

    def test_assigned_course_ids(self, app, seed):
        docente = db.session.get(User, seed["docente_id"])
        ids = docente.assigned_course_ids()
        assert seed["course_id"] in ids

    def test_repr(self, app, seed):
        admin = db.session.get(User, seed["admin_id"])
        assert "admin" in repr(admin)


class TestStudentModel:
    def test_create_student(self, app, seed):
        s = db.session.get(Student, seed["student_id"])
        assert s.codigo == "2026-001"
        assert s.nivel == "PRIMARIA"

    def test_full_name(self, app, seed):
        s = db.session.get(Student, seed["student_id"])
        assert s.full_name == "PEREZ GOMEZ, Juan"

    def test_aula(self, app, seed):
        s = db.session.get(Student, seed["student_id"])
        assert s.aula == "3 A"


class TestAcademicModels:
    def test_term_creation(self, app, seed):
        t = db.session.get(Term, seed["term_id"])
        assert t.nombre == "I Bimestre"
        assert t.anio == 2026

    def test_course_creation(self, app, seed):
        c = db.session.get(Course, seed["course_id"])
        assert c.area == "Matemática"

    def test_grade_unique_constraint(self, app, seed):
        g1 = Grade(student_id=seed["student_id"], course_id=seed["course_id"],
                    term_id=seed["term_id"], numeric_value=15)
        db.session.add(g1)
        db.session.commit()

        g2 = Grade(student_id=seed["student_id"], course_id=seed["course_id"],
                    term_id=seed["term_id"], numeric_value=18)
        db.session.add(g2)
        with pytest.raises(Exception):
            db.session.commit()
        db.session.rollback()

    def test_attendance_model(self, app, seed):
        a = Attendance(student_id=seed["student_id"], mes="Marzo", anio=2026, faltas=2, tardanzas=1)
        db.session.add(a)
        db.session.commit()
        assert a.id is not None

    def test_behavior_qualitative(self, app, seed):
        b = Behavior(student_id=seed["student_id"], indicador="Respeto",
                     eda_id=seed["eda_id"], calificacion=19)
        db.session.add(b)
        db.session.commit()
        assert b.qualitative_grade == "AD"

    def test_eda_grade(self, app, seed):
        eg = EdaGrade(student_id=seed["student_id"], course_id=seed["course_id"],
                      eda_id=seed["eda_id"], numeric_value=16)
        db.session.add(eg)
        db.session.commit()
        assert eg.qualitative_grade == "A"

    def test_registro_semana_promedio(self, app, seed):
        rs = RegistroSemana(
            student_id=seed["student_id"], course_id=seed["course_id"],
            eda_id=seed["eda_id"], semana=1,
            tarea=14, intervencion=16, fast_test=12, aptitudinal=18,
        )
        db.session.add(rs)
        db.session.commit()
        assert rs.promedio == 15

    def test_registro_semana3_includes_extras(self, app, seed):
        rs = RegistroSemana(
            student_id=seed["student_id"], course_id=seed["course_id"],
            eda_id=seed["eda_id"], semana=3,
            tarea=10, intervencion=10, fast_test=10, aptitudinal=10,
            rev_cuaderno=20, rev_libro=20,
        )
        db.session.add(rs)
        db.session.commit()
        # (10+10+10+10+20+20)/6 = 13.33 → 13
        assert rs.promedio == 13

    def test_registro_semana_promedio_none(self, app, seed):
        rs = RegistroSemana(
            student_id=seed["student_id"], course_id=seed["course_id"],
            eda_id=seed["eda_id"], semana=2,
        )
        db.session.add(rs)
        db.session.commit()
        assert rs.promedio is None

    def test_parent_responsibility_qualitative(self, app, seed):
        pr = ParentResponsibility(
            student_id=seed["student_id"], indicador="Reuniones",
            term_id=seed["term_id"], calificacion=12,
        )
        db.session.add(pr)
        db.session.commit()
        assert pr.qualitative_grade == "B"


# ─── 3. Tests del servicio de notas ─────────────────────────────────

class TestGradeService:
    def test_qualitative_ad(self):
        assert numeric_to_qualitative(20, "PRIMARIA") == "AD"
        assert numeric_to_qualitative(18, "SECUNDARIA") == "AD"

    def test_qualitative_a(self):
        assert numeric_to_qualitative(15, "PRIMARIA") == "A"
        assert numeric_to_qualitative(14, "PRIMARIA") == "A"

    def test_qualitative_b(self):
        assert numeric_to_qualitative(13, "PRIMARIA") == "B"
        assert numeric_to_qualitative(11, "PRIMARIA") == "B"

    def test_qualitative_c(self):
        assert numeric_to_qualitative(10, "PRIMARIA") == "C"
        assert numeric_to_qualitative(0, "PRIMARIA") == "C"

    def test_qualitative_none(self):
        assert numeric_to_qualitative(None) == "--"

    def test_qualitative_inicial_no_ad(self):
        assert numeric_to_qualitative(20, "INICIAL") == "A"
        assert numeric_to_qualitative(14, "INICIAL") == "A"
        assert numeric_to_qualitative(13, "INICIAL") == "B"
        assert numeric_to_qualitative(10, "INICIAL") == "C"

    def test_upsert_grade_create(self, app, seed):
        g = upsert_grade(seed["student_id"], seed["course_id"], seed["term_id"], 17)
        assert g.id is not None
        assert g.numeric_value == 17

    def test_upsert_grade_update(self, app, seed):
        upsert_grade(seed["student_id"], seed["course_id"], seed["term_id"], 14)
        g = upsert_grade(seed["student_id"], seed["course_id"], seed["term_id"], 19)
        assert g.numeric_value == 19
        count = Grade.query.filter_by(
            student_id=seed["student_id"], course_id=seed["course_id"]
        ).count()
        assert count == 1

    def test_upsert_grade_invalid_range(self, app, seed):
        with pytest.raises(ValueError):
            upsert_grade(seed["student_id"], seed["course_id"], seed["term_id"], 25)
        with pytest.raises(ValueError):
            upsert_grade(seed["student_id"], seed["course_id"], seed["term_id"], -1)

    def test_upsert_grade_none(self, app, seed):
        g = upsert_grade(seed["student_id"], seed["course_id"], seed["term_id"], None)
        assert g.numeric_value is None

    def test_grade_qualitative_property(self, app, seed):
        g = upsert_grade(seed["student_id"], seed["course_id"], seed["term_id"], 18)
        assert g.qualitative_grade == "AD"


# ─── 4. Tests de autenticación ──────────────────────────────────────

class TestAuth:
    def test_login_page_loads(self, client, seed):
        resp = client.get("/auth/login")
        assert resp.status_code == 200

    def test_login_valid(self, client, seed):
        resp = login(client, "admin", "admin1234")
        assert resp.status_code == 200

    def test_login_invalid_password(self, client, seed):
        resp = login(client, "admin", "wrongpass")
        assert "incorrectos" in resp.text

    def test_login_invalid_user(self, client, seed):
        resp = login(client, "noexiste", "pass")
        assert "incorrectos" in resp.text

    def test_logout(self, client, seed):
        login(client, "admin", "admin1234")
        resp = client.get("/auth/logout", follow_redirects=True)
        assert resp.status_code == 200

    def test_redirect_when_not_logged(self, client, seed):
        resp = client.get("/students/", follow_redirects=False)
        assert resp.status_code == 303


# ─── 5. Tests de permisos por rol ───────────────────────────────────

class TestPermissions:
    def test_admin_access_students(self, client, seed):
        login(client, "admin", "admin1234")
        resp = client.get("/students/")
        assert resp.status_code == 200

    def test_admin_access_admin_panel(self, client, seed):
        login(client, "admin", "admin1234")
        resp = client.get("/admin/users")
        assert resp.status_code == 200

    def test_docente_cannot_access_admin(self, client, seed):
        login(client, "docente", "docente1234")
        resp = client.get("/admin/users", follow_redirects=False)
        assert resp.status_code in (303, 403)

    def test_auxiliar_cannot_access_grades(self, client, seed):
        login(client, "auxiliar", "auxiliar1234")
        resp = client.get("/grades/matrix", follow_redirects=False)
        assert resp.status_code in (303, 403)


class TestEdaMatrixFeatureFlag:
    def test_docente_redirected_when_eda_matrix_disabled(self, client, seed):
        from app.services.feature_flags import set_eda_matrix_enabled_for_docentes

        set_eda_matrix_enabled_for_docentes(False)
        try:
            login(client, "docente", "docente1234")
            resp = client.get("/grades/eda-matrix", follow_redirects=False)
            assert resp.status_code == 303
            assert "/dashboard" in (resp.headers.get("location") or "")
        finally:
            set_eda_matrix_enabled_for_docentes(True)

    def test_admin_eda_matrix_ok_when_disabled_for_docentes(self, client, seed):
        from app.services.feature_flags import set_eda_matrix_enabled_for_docentes

        set_eda_matrix_enabled_for_docentes(False)
        try:
            login(client, "admin", "admin1234")
            resp = client.get("/grades/eda-matrix", follow_redirects=True)
            assert resp.status_code == 200
        finally:
            set_eda_matrix_enabled_for_docentes(True)

    def test_admin_feature_flags_page(self, client, seed):
        login(client, "admin", "admin1234")
        resp = client.get("/admin/feature-flags")
        assert resp.status_code == 200
        assert "Matriz de EDAs" in resp.text


# ─── 6. Tests de rutas principales ──────────────────────────────────

class TestRoutes:
    def test_dashboard(self, client, seed):
        login(client, "admin", "admin1234")
        resp = client.get("/dashboard")
        assert resp.status_code == 200

    def test_students_list(self, client, seed):
        login(client, "admin", "admin1234")
        resp = client.get("/students/")
        assert resp.status_code == 200

    def test_grades_page(self, client, seed):
        login(client, "admin", "admin1234")
        resp = client.get("/grades/matrix")
        assert resp.status_code == 200

    def test_attendance_page(self, client, seed):
        login(client, "admin", "admin1234")
        resp = client.get("/attendance/")
        assert resp.status_code == 200

    def test_behavior_page(self, client, seed):
        login(client, "admin", "admin1234")
        resp = client.get("/behavior/")
        assert resp.status_code == 200

    def test_404_page(self, client, seed):
        login(client, "admin", "admin1234")
        resp = client.get("/ruta-inexistente")
        assert resp.status_code == 404


# ─── 7. Tests de cascada y relaciones ───────────────────────────────

class TestCascade:
    def test_delete_student_cascades_grades(self, app, seed):
        upsert_grade(seed["student_id"], seed["course_id"], seed["term_id"], 15)
        assert Grade.query.count() == 1
        student = db.session.get(Student, seed["student_id"])
        db.session.delete(student)
        db.session.commit()
        assert Grade.query.count() == 0

    def test_delete_student_cascades_attendance(self, app, seed):
        a = Attendance(student_id=seed["student_id"], mes="Abril", anio=2026, faltas=0, tardanzas=0)
        db.session.add(a)
        db.session.commit()
        student = db.session.get(Student, seed["student_id"])
        db.session.delete(student)
        db.session.commit()
        assert Attendance.query.count() == 0

    def test_delete_student_cascades_behavior(self, app, seed):
        b = Behavior(student_id=seed["student_id"], indicador="Orden",
                     eda_id=seed["eda_id"], calificacion=15)
        db.session.add(b)
        db.session.commit()
        student = db.session.get(Student, seed["student_id"])
        db.session.delete(student)
        db.session.commit()
        assert Behavior.query.count() == 0


# ─── 8. Tests de edge cases ─────────────────────────────────────────

class TestEdgeCases:
    def test_qualitative_boundary_values(self):
        assert numeric_to_qualitative(18, "PRIMARIA") == "AD"
        assert numeric_to_qualitative(17, "PRIMARIA") == "A"
        assert numeric_to_qualitative(14, "PRIMARIA") == "A"
        assert numeric_to_qualitative(13, "PRIMARIA") == "B"
        assert numeric_to_qualitative(11, "PRIMARIA") == "B"
        assert numeric_to_qualitative(10, "PRIMARIA") == "C"

    def test_behavior_qualitative_all_ranges(self, app, seed):
        cases = [(20, "AD"), (18, "AD"), (17, "A"), (15, "A"), (14, "B"), (11, "B"), (10, "C"), (0, "C")]
        for val, expected in cases:
            b = Behavior(student_id=seed["student_id"], indicador=f"Test{val}",
                         eda_id=seed["eda_id"], calificacion=val)
            assert b.qualitative_grade == expected

    def test_behavior_qualitative_none(self, app, seed):
        b = Behavior(student_id=seed["student_id"], indicador="X",
                     eda_id=seed["eda_id"], calificacion=None)
        assert b.qualitative_grade is None

    def test_user_inactive_cannot_login(self, app, client, seed):
        docente = db.session.get(User, seed["docente_id"])
        docente.is_active = False
        db.session.commit()
        resp = login(client, "docente", "docente1234")
        assert "incorrectos" in resp.text


# ─── 9. Endurecimiento (cabeceras, CSRF, redirects, IDOR) ─────────

class TestSecurityHardening:
    def test_security_headers_on_login_page(self, client):
        r = client.get("/auth/login")
        assert r.status_code == 200
        assert r.headers.get("x-frame-options", "").upper() == "DENY"
        assert "nosniff" in (r.headers.get("x-content-type-options") or "").lower()
        csp = r.headers.get("content-security-policy") or ""
        assert csp
        assert "cdn.jsdelivr.net" in csp  # Chart.js en dashboard / consolidado
        assert "strict-origin-when-cross-origin" in (r.headers.get("referrer-policy") or "").lower()

    def test_login_post_rejects_mismatched_csrf_header(self, client, seed):
        page = client.get("/auth/login")
        m = re.search(r'name="csrf_token"\s+value="([^"]+)"', page.text)
        token = m.group(1) if m else ""
        # Sin cabecera CSRF válida y con Origin ajeno al host, no aplica el bypass same-site.
        resp = client.post(
            "/auth/login",
            data={"username": "admin", "password": "admin1234", "csrf_token": token},
            headers={
                "X-CSRFToken": "token-invalido-forzado",
                "Origin": "https://evil.example",
            },
        )
        assert resp.status_code == 403

    def test_login_redirect_ignores_external_next(self, client, seed):
        page = client.get("/auth/login")
        m = re.search(r'name="csrf_token"\s+value="([^"]+)"', page.text)
        token = m.group(1) if m else ""
        resp = client.post(
            "/auth/login?next=https://evil.example/phish",
            data={"username": "admin", "password": "admin1234", "csrf_token": token},
            headers={"X-CSRFToken": token},
            follow_redirects=False,
        )
        assert resp.status_code == 303
        loc = resp.headers.get("location", "")
        assert loc.startswith("/")
        assert "evil.example" not in loc

    def test_docente_idor_cannot_open_secundaria_student_grades(self, app, client, seed):
        from app.utils.id_mask import encode_id

        s2 = Student(
            codigo="2026-SEC-01",
            nombres="Ana",
            apellido_paterno="LOPEZ",
            apellido_materno="RUIZ",
            nivel="SECUNDARIA",
            grado="1",
            seccion="A",
            estado="ACTIVO",
        )
        db.session.add(s2)
        db.session.commit()

        login(client, "docente", "docente1234")
        token = encode_id(s2.id)
        resp = client.get(f"/grades/student/{token}")
        assert resp.status_code == 403


# ─── 10. Helpers: generador de Excel en memoria ─────────────────────

def _make_teachers_xlsx(rows):
    """
    Crea un .xlsx en BytesIO a partir de una lista de listas.
    Primera lista = cabecera, siguientes = filas de datos.
    """
    import openpyxl
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _xlsx_bytes(rows) -> bytes:
    return _make_teachers_xlsx(rows).read()


# ─── 11. Tests de funciones puras del servicio ──────────────────────

class TestCellToStr:
    def test_none(self):
        assert _cell_to_str(None) == ""

    def test_string(self):
        assert _cell_to_str("  hello ") == "hello"

    def test_int(self):
        assert _cell_to_str(3) == "3"

    def test_float_whole(self):
        assert _cell_to_str(3.0) == "3"

    def test_float_fractional(self):
        result = _cell_to_str(3.5)
        assert "3.5" in result

    def test_bool_true(self):
        assert _cell_to_str(True) in ("True", "1")

    def test_bool_false(self):
        assert _cell_to_str(False) in ("False", "0")


class TestParseCourseIds:
    def test_empty_string(self):
        ids, bad = _parse_course_ids("")
        assert ids == []
        assert bad == []

    def test_none(self):
        ids, bad = _parse_course_ids(None)
        assert ids == []

    def test_single_id(self):
        ids, bad = _parse_course_ids("5")
        assert ids == [5]
        assert bad == []

    def test_comma_separated(self):
        ids, _ = _parse_course_ids("1,2,3")
        assert ids == [1, 2, 3]

    def test_semicolon_separated(self):
        ids, _ = _parse_course_ids("1;2;3")
        assert ids == [1, 2, 3]

    def test_spaces(self):
        ids, _ = _parse_course_ids("  1 , 2 , 3  ")
        assert ids == [1, 2, 3]

    def test_whitespace_only(self):
        ids, _ = _parse_course_ids("   ")
        assert ids == []

    def test_float_as_string_reported(self):
        ids, bad = _parse_course_ids("5.0")
        assert bad == ["5.0"], "Decimal notation is not a valid integer"

    def test_duplicate_ids_deduplicated(self):
        """Duplicate IDs must be deduplicated to avoid IntegrityError
        on TeacherCourse unique constraint."""
        ids, _ = _parse_course_ids("1,1,2")
        assert ids == [1, 2]

    def test_non_numeric_reported(self):
        """Non-numeric tokens must be returned in the bad list, not silently dropped."""
        ids, bad = _parse_course_ids("1,abc,3")
        assert ids == [1, 3]
        assert bad == ["abc"]


class TestGradosValidos:
    def test_inicial(self):
        assert _grados_validos("INICIAL") == {"3", "4", "5"}

    def test_primaria(self):
        assert _grados_validos("PRIMARIA") == {"1", "2", "3", "4", "5", "6"}

    def test_secundaria(self):
        assert _grados_validos("SECUNDARIA") == {"1", "2", "3", "4", "5"}

    def test_unknown(self):
        assert _grados_validos("JARDIN") == set()


# ─── 12. Tests del servicio de importación de docentes ───────────────

class TestImportTeachersService:

    def test_happy_path_minimal(self, app, seed):
        """Importar un docente con solo columnas obligatorias."""
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO"],
            ["profe.test", "Profesor de Prueba"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 1
        assert len(result["errores"]) == 0
        user = User.query.filter_by(username="profe.test").first()
        assert user is not None
        assert user.role == RoleEnum.DOCENTE
        assert user.is_active is True

    def test_happy_path_full_columns(self, app, seed):
        """Importar docente con todas las columnas, incluidos cursos."""
        cid = seed["course_id"]
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO", "PASSWORD", "NIVEL", "GRADO", "IDS_CURSOS"],
            ["profe.full", "Prof Completo", "MiClave123", "PRIMARIA", "3", str(cid)],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 1
        assert len(result["errores"]) == 0
        user = User.query.filter_by(username="profe.full").first()
        assert user.check_password("MiClave123")
        assert user.nivel == "PRIMARIA"
        assert user.grado == "3"
        tc = TeacherCourse.query.filter_by(user_id=user.id, course_id=cid).first()
        assert tc is not None

    def test_password_auto_generated(self, app, seed):
        """Si PASSWORD está vacío, se genera automáticamente."""
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO", "PASSWORD"],
            ["profe.auto", "Prof Auto", ""],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 1
        detalle = result["detalle"][0]
        assert detalle["password_es_temporal"] is True
        pwd = detalle["password_nota"]
        user = User.query.filter_by(username="profe.auto").first()
        assert user.check_password(pwd), "Auto-generated password must work"

    def test_password_from_file(self, app, seed):
        """Si PASSWORD tiene valor, se usa esa y no se muestra en detalle."""
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO", "PASSWORD"],
            ["profe.pwd", "Prof Pwd", "SecretPass99"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 1
        detalle = result["detalle"][0]
        assert detalle["password_es_temporal"] is False
        assert detalle["password_nota"] == "La indicada en el Excel"

    def test_multiple_rows(self, app, seed):
        """Importar varios docentes de una vez."""
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO"],
            ["profe.a", "Profesor A"],
            ["profe.b", "Profesor B"],
            ["profe.c", "Profesor C"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 3
        assert len(result["errores"]) == 0

    def test_empty_rows_skipped(self, app, seed):
        """Filas vacías no generan errores."""
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO"],
            ["profe.ok", "Profesor OK"],
            [None, None],
            ["", ""],
            ["profe.ok2", "Profesor OK2"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 2
        assert len(result["errores"]) == 0

    def test_missing_required_columns_raises(self, app, seed):
        """Excel sin columnas obligatorias lanza ValueError."""
        buf = _make_teachers_xlsx([
            ["FOO", "BAR"],
            ["x", "y"],
        ])
        with pytest.raises(ValueError, match="Faltan columnas obligatorias"):
            import_teachers_from_excel(buf)

    def test_empty_excel_raises(self, app, seed):
        buf = _make_teachers_xlsx([])
        with pytest.raises(ValueError, match="vacío"):
            import_teachers_from_excel(buf)

    def test_usuario_vacio(self, app, seed):
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO"],
            ["", "Prof sin usuario"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 0
        assert len(result["errores"]) == 1
        assert "vacío" in result["errores"][0]["motivo"].lower()

    def test_nombre_vacio(self, app, seed):
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO"],
            ["profe.test", ""],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 0
        assert len(result["errores"]) == 1
        assert "NOMBRE_COMPLETO" in result["errores"][0]["motivo"]

    def test_invalid_username_short(self, app, seed):
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO"],
            ["ab", "Prof Corto"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 0
        assert len(result["errores"]) == 1
        assert "inválido" in result["errores"][0]["motivo"].lower()

    def test_invalid_username_special_chars(self, app, seed):
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO"],
            ["profe @#$", "Prof Especial"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 0
        assert len(result["errores"]) == 1

    def test_duplicate_username_in_db(self, app, seed):
        """Username que ya existe en BD genera error."""
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO"],
            ["docente", "Duplicado del docente seed"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 0
        assert len(result["errores"]) == 1
        assert "ya existe" in result["errores"][0]["motivo"]

    def test_duplicate_username_within_same_file(self, app, seed):
        """Mismo username en dos filas: primera OK, segunda error."""
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO"],
            ["profe.dup", "Primera Vez"],
            ["profe.dup", "Segunda Vez"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 1
        assert len(result["errores"]) == 1
        assert "ya existe" in result["errores"][0]["motivo"]

    def test_invalid_nivel(self, app, seed):
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO", "NIVEL"],
            ["profe.mal", "Prof Mal Nivel", "UNIVERSIDAD"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 0
        assert "inválido" in result["errores"][0]["motivo"].lower()

    def test_grado_without_nivel_rejected(self, app, seed):
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO", "NIVEL", "GRADO"],
            ["profe.nogr", "Prof Solo Grado", "", "3"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 0
        assert "NIVEL" in result["errores"][0]["motivo"]

    def test_invalid_grado_for_nivel(self, app, seed):
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO", "NIVEL", "GRADO"],
            ["profe.bad", "Prof Grado 6 Inicial", "INICIAL", "6"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 0
        assert "GRADO" in result["errores"][0]["motivo"]

    def test_nivel_without_grado_is_ok(self, app, seed):
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO", "NIVEL", "GRADO"],
            ["profe.niv", "Prof Solo Nivel", "SECUNDARIA", ""],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 1
        user = User.query.filter_by(username="profe.niv").first()
        assert user.nivel == "SECUNDARIA"
        assert user.grado is None

    def test_nonexistent_course_ids(self, app, seed):
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO", "IDS_CURSOS"],
            ["profe.bad", "Prof Bad Course", "99999"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 0
        assert "inexistentes" in result["errores"][0]["motivo"]

    def test_mixed_valid_and_invalid_rows(self, app, seed):
        """Verifica que filas válidas se importan y las inválidas generan errores,
        sin que un error en fila N corrompa la fila N+1."""
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO"],
            ["profe.ok1", "Profesor Bueno 1"],
            ["ab", "Nombre Corto Usuario"],
            ["profe.ok2", "Profesor Bueno 2"],
            ["docente", "Duplicado seed"],
            ["profe.ok3", "Profesor Bueno 3"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 3
        assert len(result["errores"]) == 2

    def test_error_after_duplicate_does_not_corrupt_session(self, app, seed):
        """BUG FINDER: after IntegrityError + rollback for a duplicate username,
        subsequent valid rows must still be importable. A broken session would
        cause all remaining rows to fail."""
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO"],
            ["docente", "Duplicado del seed"],
            ["profe.afterdup", "Profesor Post-Dup"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 1, "Row after IntegrityError must succeed"
        assert len(result["errores"]) == 1

    # ── BUG: IDs de curso duplicados ──────────────────────────────────

    def test_duplicate_course_ids_in_excel(self, app, seed):
        """BUG FINDER: IDS_CURSOS = '1,1' with real course ID should NOT
        cause IntegrityError on uq_teacher_course."""
        cid = seed["course_id"]
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO", "IDS_CURSOS"],
            ["profe.dups", "Prof Dup Cursos", f"{cid},{cid}"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 1, (
            "Duplicate course IDs should be deduplicated, not crash"
        )
        assert len(result["errores"]) == 0
        tc_count = TeacherCourse.query.filter_by(
            user_id=User.query.filter_by(username="profe.dups").first().id
        ).count()
        assert tc_count == 1, "Only one TeacherCourse per course"

    # ── BUG: IDs no numéricos se descartan sin avisar ────────────────

    def test_non_numeric_course_ids_reported_as_error(self, app, seed):
        """IDS_CURSOS with garbage like 'abc' must produce an error row,
        not silently drop the invalid token."""
        cid = seed["course_id"]
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO", "IDS_CURSOS"],
            ["profe.bad", "Prof Bad Ids", f"{cid},abc"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 0, "Row with invalid course IDs should not be imported"
        assert len(result["errores"]) == 1
        assert "abc" in result["errores"][0]["motivo"]

    # ── Grado como número en Excel (openpyxl devuelve int) ───────────

    def test_grado_as_integer_from_excel(self, app, seed):
        """openpyxl devuelve celdas numéricas como int/float, no str.
        El grado '3' llega como int 3 — debe convertirse correctamente."""
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO", "NIVEL", "GRADO"],
            ["profe.int", "Prof Int Grado", "PRIMARIA", 3],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 1, f"Errors: {result['errores']}"
        user = User.query.filter_by(username="profe.int").first()
        assert user.grado == "3"

    def test_course_ids_as_integer_from_excel(self, app, seed):
        """openpyxl puede devolver IDS_CURSOS como int si solo hay un ID."""
        cid = seed["course_id"]
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE_COMPLETO", "IDS_CURSOS"],
            ["profe.intid", "Prof Int ID", cid],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 1, f"Errors: {result['errores']}"

    # ── Cabeceras case-insensitive y con espacios ─────────────────────

    def test_headers_case_insensitive(self, app, seed):
        buf = _make_teachers_xlsx([
            ["usuario", "nombre_completo"],
            ["profe.low", "Prof Lower Case Headers"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 1

    def test_headers_with_spaces(self, app, seed):
        """Cabecera 'NOMBRE COMPLETO' (con espacio) se normaliza a NOMBRE_COMPLETO."""
        buf = _make_teachers_xlsx([
            ["USUARIO", "NOMBRE COMPLETO"],
            ["profe.spc", "Prof Spaces"],
        ])
        result = import_teachers_from_excel(buf)
        assert result["creados"] == 1

    # ── Template Excel generation ─────────────────────────────────────

    def test_template_generates_valid_xlsx(self, app, seed):
        buf = generate_teachers_template_excel()
        assert buf.read(2) == b"PK", "Must be valid .xlsx (ZIP)"

    def test_template_has_correct_sheets(self, app, seed):
        import openpyxl
        buf = generate_teachers_template_excel()
        wb = openpyxl.load_workbook(buf)
        names = wb.sheetnames
        assert "Docentes" in names
        assert "Cursos_ID" in names
        assert "Instrucciones" in names

    def test_template_headers_match_service_columns(self, app, seed):
        import openpyxl
        from app.services.excel_import_teachers import COLUMNAS
        buf = generate_teachers_template_excel()
        wb = openpyxl.load_workbook(buf)
        ws = wb["Docentes"]
        headers = [cell.value for cell in ws[1]]
        assert headers == COLUMNAS

    def test_template_catalog_includes_courses(self, app, seed):
        """La hoja Cursos_ID debe listar los cursos del sistema."""
        import openpyxl
        buf = generate_teachers_template_excel()
        wb = openpyxl.load_workbook(buf)
        ws = wb["Cursos_ID"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        course_ids = [r[0] for r in rows]
        assert seed["course_id"] in course_ids


# ─── 13. Tests HTTP de rutas de importación de docentes ──────────────

class TestTeacherImportRoutes:

    def _csrf_headers(self, client):
        """Get CSRF token from users page (must be logged in first)."""
        page = client.get("/admin/users")
        m = re.search(r'name="csrf_token"\s+value="([^"]+)"', page.text)
        token = m.group(1) if m else ""
        return {"X-CSRFToken": token}

    def test_users_page_shows_import_section(self, client, seed):
        login(client, "admin", "admin1234")
        resp = client.get("/admin/users")
        assert resp.status_code == 200
        assert "importar-docentes" in resp.text.lower() or "Importar docentes" in resp.text

    def test_download_template_requires_admin(self, client, seed):
        login(client, "docente", "docente1234")
        resp = client.get("/admin/users/plantilla-docentes", follow_redirects=False)
        assert resp.status_code in (303, 403)

    def test_download_template_as_admin(self, client, seed):
        login(client, "admin", "admin1234")
        resp = client.get("/admin/users/plantilla-docentes")
        assert resp.status_code == 200
        ct = resp.headers.get("content-type", "")
        assert "spreadsheetml" in ct or "application/" in ct
        assert resp.content[:2] == b"PK"

    def test_import_requires_admin(self, client, seed):
        login(client, "docente", "docente1234")
        resp = client.post(
            "/admin/users/importar-docentes",
            files={"excel_file": ("test.xlsx", b"PK fake", "application/octet-stream")},
            follow_redirects=False,
        )
        assert resp.status_code in (303, 403)

    def test_import_rejects_non_xlsx(self, client, seed):
        login(client, "admin", "admin1234")
        hdrs = self._csrf_headers(client)
        resp = client.post(
            "/admin/users/importar-docentes",
            files={"excel_file": ("test.csv", b"a,b,c", "text/csv")},
            headers=hdrs,
            follow_redirects=True,
        )
        assert resp.status_code == 200
        assert ".xlsx" in resp.text

    def test_import_rejects_fake_xlsx(self, client, seed):
        """File with .xlsx extension but non-ZIP content."""
        login(client, "admin", "admin1234")
        hdrs = self._csrf_headers(client)
        resp = client.post(
            "/admin/users/importar-docentes",
            files={"excel_file": ("test.xlsx", b"NOT A ZIP FILE", "application/octet-stream")},
            headers=hdrs,
            follow_redirects=True,
        )
        assert resp.status_code == 200
        assert "no parece" in resp.text or "válido" in resp.text

    def test_import_valid_xlsx_creates_users(self, client, seed):
        login(client, "admin", "admin1234")
        hdrs = self._csrf_headers(client)
        xlsx = _xlsx_bytes([
            ["USUARIO", "NOMBRE_COMPLETO"],
            ["profe.http", "Prof HTTP Test"],
        ])
        resp = client.post(
            "/admin/users/importar-docentes",
            files={"excel_file": ("docentes.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=hdrs,
            follow_redirects=True,
        )
        assert resp.status_code == 200
        assert "Resultado" in resp.text or "creados" in resp.text.lower() or "Docentes creados" in resp.text
        user = User.query.filter_by(username="profe.http").first()
        assert user is not None

    def test_import_no_file_redirects(self, client, seed):
        login(client, "admin", "admin1234")
        hdrs = self._csrf_headers(client)
        resp = client.post(
            "/admin/users/importar-docentes",
            data={},
            headers=hdrs,
            follow_redirects=True,
        )
        assert resp.status_code == 200
