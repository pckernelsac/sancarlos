"""
Importación masiva de estudiantes desde Excel (.xlsx).
Columnas esperadas (cabecera en fila 1, datos desde fila 2):
  APELLIDO_PATERNO | APELLIDO_MATERNO | NOMBRES | NIVEL | GRADO | SECCION | DNI | FECHA_NACIMIENTO | ESTADO
El CODIGO se genera automáticamente por el sistema.
"""
import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import db
from app.models.student import Student, GRADOS, SECCIONES, ESTADOS, NIVELES
from app.services.student_service import generate_student_code

COLUMNAS = ["APELLIDO_PATERNO", "APELLIDO_MATERNO", "NOMBRES", "NIVEL", "GRADO", "SECCION", "DNI", "FECHA_NACIMIENTO", "ESTADO"]
REQUERIDAS = {"APELLIDO_PATERNO", "NOMBRES", "GRADO", "SECCION"}


# ── Lector / importador ────────────────────────────────────────────────────────

def import_students_from_excel(file_stream) -> dict:
    """
    Procesa un archivo Excel y retorna:
    {
      "insertados": int,
      "actualizados": int,
      "omitidos": int,
      "errores": [ {"fila": int, "codigo": str, "motivo": str} ],
      "detalle": [ {"fila": int, "codigo": str, "nombre": str, "accion": str} ]
    }
    """
    try:
        wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise ValueError("El archivo no es un Excel válido (.xlsx).")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El archivo está vacío.")

    # Detecta cabecera — primera fila normalizada
    header = [str(c).strip().upper() if c else "" for c in rows[0]]

    # Mapea nombre de columna → índice
    col_idx = {name: header.index(name) for name in COLUMNAS if name in header}

    missing = REQUERIDAS - col_idx.keys()
    if missing:
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(sorted(missing))}")

    result = {"insertados": 0, "actualizados": 0, "omitidos": 0, "errores": [], "detalle": []}

    for row_num, row in enumerate(rows[1:], start=2):
        def get(col):
            idx = col_idx.get(col)
            val = row[idx] if idx is not None and idx < len(row) else None
            return str(val).strip() if val not in (None, "") else ""

        ap_paterno = get("APELLIDO_PATERNO").upper()
        ap_materno = get("APELLIDO_MATERNO").upper()
        nombres    = get("NOMBRES").upper()
        nivel      = get("NIVEL").upper() or "PRIMARIA"
        grado      = get("GRADO")
        seccion    = get("SECCION").upper()
        dni        = get("DNI") or None
        estado     = get("ESTADO").upper() or "ACTIVO"
        fecha_str  = get("FECHA_NACIMIENTO")

        # Fila completamente vacía → saltar silenciosamente
        if not any([ap_paterno, nombres, grado, seccion]):
            continue

        # Validaciones
        errores_fila = []
        if not ap_paterno:
            errores_fila.append("APELLIDO_PATERNO vacío")
        if not nombres:
            errores_fila.append("NOMBRES vacío")
        if nivel not in NIVELES:
            nivel = "PRIMARIA"
        if grado not in GRADOS:
            errores_fila.append(f"GRADO '{grado}' no válido (usar: {', '.join(GRADOS)})")
        if seccion not in SECCIONES:
            errores_fila.append(f"SECCION '{seccion}' no válida (usar: {', '.join(SECCIONES)})")
        if estado not in ESTADOS:
            estado = "ACTIVO"

        # Fecha de nacimiento (acepta DD/MM/YYYY o YYYY-MM-DD)
        fecha = None
        if fecha_str:
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    fecha = datetime.datetime.strptime(fecha_str, fmt).date()
                    break
                except ValueError:
                    continue
            if fecha is None:
                errores_fila.append(f"FECHA_NACIMIENTO '{fecha_str}' no reconocida (usar DD/MM/YYYY)")

        if errores_fila:
            result["errores"].append({
                "fila": row_num,
                "codigo": "—",
                "motivo": " | ".join(errores_fila),
            })
            continue

        nombre_completo = f"{ap_paterno} {ap_materno}, {nombres}".strip()

        # Verificar duplicados: primero por DNI, luego por nombre+nivel+grado
        existing = None
        if dni:
            existing = Student.query.filter_by(dni=dni).first()
        if not existing:
            existing = Student.query.filter_by(
                apellido_paterno=ap_paterno,
                apellido_materno=ap_materno,
                nombres=nombres,
                nivel=nivel,
                grado=grado,
            ).first()

        if existing:
            existing.apellido_paterno = ap_paterno
            existing.apellido_materno = ap_materno
            existing.nombres          = nombres
            existing.nivel            = nivel
            existing.grado            = grado
            existing.seccion          = seccion
            existing.estado           = estado
            existing.fecha_nacimiento = fecha
            if dni:
                existing.dni = dni
            result["actualizados"] += 1
            result["detalle"].append({
                "fila": row_num,
                "codigo": existing.codigo,
                "nombre": nombre_completo,
                "accion": "ACTUALIZADO",
            })
            continue

        # Generar código único automáticamente
        codigo = generate_student_code(ap_paterno, ap_materno)

        db.session.add(Student(
            codigo=codigo, apellido_paterno=ap_paterno, apellido_materno=ap_materno,
            nombres=nombres, nivel=nivel, grado=grado, seccion=seccion, dni=dni,
            estado=estado, fecha_nacimiento=fecha,
        ))
        db.session.flush()
        result["insertados"] += 1

        result["detalle"].append({
            "fila": row_num,
            "codigo": codigo,
            "nombre": nombre_completo,
            "accion": "INSERTADO",
        })

    db.session.commit()
    wb.close()
    return result


# ── Generador de plantilla ─────────────────────────────────────────────────────

def generate_template_excel() -> BytesIO:
    """Genera un .xlsx de plantilla con cabecera, ejemplo y notas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    # Estilos
    header_fill = PatternFill("solid", fgColor="0E2F77")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    example_fill = PatternFill("solid", fgColor="EEF2FF")
    example_font = Font(color="374151", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # Cabecera
    ws.append(COLUMNAS)
    for col_num, _ in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_fill and header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Fila de ejemplo
    ejemplo = [
        "GARCIA", "LOPEZ", "ANA LUCIA",
        "PRIMARIA", "1", "A", "12345678", "15/03/2018", "ACTIVO"
    ]
    ws.append(ejemplo)
    for col_num in range(1, len(COLUMNAS) + 1):
        cell = ws.cell(row=2, column=col_num)
        cell.fill = example_fill
        cell.font = example_font
        cell.alignment = center
        cell.border = border

    # Ajuste de ancho de columnas
    col_widths = [20, 20, 22, 14, 8, 10, 12, 20, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Hoja de instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["INSTRUCCIONES DE USO"],
        [""],
        ["CAMPO",          "OBLIGATORIO", "VALORES ACEPTADOS"],
        ["APELLIDO_PATERNO","SI",          "Texto en mayúsculas"],
        ["APELLIDO_MATERNO","NO",          "Texto en mayúsculas"],
        ["NOMBRES",        "SI",          "Texto en mayúsculas"],
        ["NIVEL",          "NO",          ", ".join(NIVELES) + " (default: PRIMARIA)"],
        ["GRADO",          "SI",          ", ".join(GRADOS)],
        ["SECCION",        "SI",          ", ".join(SECCIONES)],
        ["DNI",            "NO",          "8 dígitos numéricos (si existe, actualiza el registro)"],
        ["FECHA_NACIMIENTO","NO",         "Formato DD/MM/YYYY"],
        ["ESTADO",         "NO",          ", ".join(ESTADOS) + " (default: ACTIVO)"],
        [""],
        ["NOTAS:"],
        ["- El CODIGO del estudiante se genera automaticamente por el sistema."],
        ["- Formato del codigo: iniciales de apellidos + año + correlativo (ej: GL20260001)."],
        ["- Si se proporciona DNI y ya existe un estudiante con ese DNI, se ACTUALIZA."],
        ["- Las filas con errores seran omitidas y reportadas al finalizar."],
    ]
    for row in instrucciones:
        ws2.append(row)

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 50
    ws2["A1"].font = Font(bold=True, size=13, color="0E2F77")
    ws2["A3"].font = Font(bold=True)
    ws2["B3"].font = Font(bold=True)
    ws2["C3"].font = Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
