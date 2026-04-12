"""
Importación masiva de docentes (usuarios rol DOCENTE) desde Excel (.xlsx).

Cabecera en fila 1 (nombres de columna sin distinguir mayúsculas):
  USUARIO | NOMBRE_COMPLETO | PASSWORD | NIVEL | GRADO | IDS_CURSOS

Obligatorias: USUARIO, NOMBRE_COMPLETO
PASSWORD: opcional; si está vacío se genera una contraseña temporal (se muestra en el resultado).
NIVEL / GRADO: opcionales (ámbito del docente).
IDS_CURSOS: IDs numéricos de cursos separados por coma o punto y coma (ej: 12,15,20).
"""
from __future__ import annotations

import re
import secrets
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import IntegrityError

from app.database import db
from app.models.academic import Course
from app.models.user import User, RoleEnum, TeacherCourse
from app.models.student import NIVELES, GRADOS_INICIAL, GRADOS_PRIMARIA, GRADOS_SECUNDARIA

COLUMNAS = ["USUARIO", "NOMBRE_COMPLETO", "PASSWORD", "NIVEL", "GRADO", "IDS_CURSOS"]
REQUERIDAS = {"USUARIO", "NOMBRE_COMPLETO"}
MAX_EXCEL_BYTES = 5 * 1024 * 1024

USER_RE = re.compile(r"^[a-zA-Z0-9_.-]{3,64}$")


def _cell_to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return str(val).strip()
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val == int(val):
            return str(int(val))
        return str(int(val)) if isinstance(val, int) else str(val).strip()
    return str(val).strip()


def _norm_header(cell) -> str:
    if cell is None:
        return ""
    return str(cell).strip().upper().replace(" ", "_")


def _grados_validos(nivel: str) -> set[str]:
    if nivel == "INICIAL":
        return set(GRADOS_INICIAL)
    if nivel == "PRIMARIA":
        return set(GRADOS_PRIMARIA)
    if nivel == "SECUNDARIA":
        return set(GRADOS_SECUNDARIA)
    return set()


def _parse_course_ids(raw: str) -> tuple[list[int], list[str]]:
    """Returns (valid_ids_deduped, invalid_tokens)."""
    if not raw or not str(raw).strip():
        return [], []
    seen: set[int] = set()
    out: list[int] = []
    bad: list[str] = []
    for part in re.split(r"[,;\s]+", str(raw).strip()):
        part = part.strip()
        if not part:
            continue
        try:
            val = int(part)
        except ValueError:
            bad.append(part)
            continue
        if val not in seen:
            seen.add(val)
            out.append(val)
    return out, bad


def import_teachers_from_excel(file_stream) -> dict:
    """
    Retorna:
      creados: int
      errores: {fila, usuario, motivo}[]
      detalle: {fila, usuario, nombre, password_nota, cursos_asignados}[]
    """
    try:
        wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        raise ValueError("El archivo no es un Excel válido (.xlsx).") from exc

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El archivo está vacío.")

    header = [_norm_header(c) for c in rows[0]]
    col_idx = {}
    for i, name in enumerate(header):
        if name and name not in col_idx:
            col_idx[name] = i

    missing = REQUERIDAS - col_idx.keys()
    if missing:
        raise ValueError(
            f"Faltan columnas obligatorias: {', '.join(sorted(missing))}. "
            f"Use: {', '.join(COLUMNAS)}."
        )

    course_ids_db = {c.id for c in Course.query.with_entities(Course.id).all()}

    result = {
        "creados": 0,
        "errores": [],
        "detalle": [],
    }

    def get_cell(row: tuple, col_name: str) -> str:
        idx = col_idx.get(col_name)
        if idx is None or idx >= len(row):
            return ""
        return _cell_to_str(row[idx])

    for row_num, row in enumerate(rows[1:], start=2):
        usuario = get_cell(row, "USUARIO")
        nombre = get_cell(row, "NOMBRE_COMPLETO")
        pwd_in = get_cell(row, "PASSWORD")
        nivel_raw = get_cell(row, "NIVEL").upper()
        nivel = nivel_raw or None
        grado = get_cell(row, "GRADO") or None
        ids_raw = get_cell(row, "IDS_CURSOS")

        if not usuario and not nombre:
            continue

        if not usuario:
            result["errores"].append({"fila": row_num, "usuario": "", "motivo": "USUARIO vacío"})
            continue
        if not nombre:
            result["errores"].append({"fila": row_num, "usuario": usuario, "motivo": "NOMBRE_COMPLETO vacío"})
            continue

        if not USER_RE.match(usuario):
            result["errores"].append({
                "fila": row_num,
                "usuario": usuario,
                "motivo": "Usuario inválido (3-64 caracteres: letras, números, . _ -)",
            })
            continue

        if grado and not nivel:
            result["errores"].append({
                "fila": row_num,
                "usuario": usuario,
                "motivo": "Indique NIVEL si escribe GRADO",
            })
            continue

        if nivel and nivel not in NIVELES:
            result["errores"].append({
                "fila": row_num,
                "usuario": usuario,
                "motivo": f"NIVEL inválido: {nivel}. Use: {', '.join(NIVELES)}",
            })
            continue

        if grado and nivel:
            gv = _grados_validos(nivel)
            if grado not in gv:
                result["errores"].append({
                    "fila": row_num,
                    "usuario": usuario,
                    "motivo": f"GRADO {grado} no válido para nivel {nivel}",
                })
                continue

        if nivel and not grado:
            grado = None

        cursos, bad_tokens = _parse_course_ids(ids_raw)
        if bad_tokens:
            result["errores"].append({
                "fila": row_num,
                "usuario": usuario,
                "motivo": f"IDS_CURSOS contiene valores no numéricos: {bad_tokens}",
            })
            continue

        bad = [cid for cid in cursos if cid not in course_ids_db]
        if bad:
            result["errores"].append({
                "fila": row_num,
                "usuario": usuario,
                "motivo": f"IDs de curso inexistentes: {bad}",
            })
            continue

        pwd_from_file = bool(pwd_in)
        password = pwd_in if pwd_from_file else secrets.token_urlsafe(12)

        user = User(
            username=usuario,
            full_name=nombre,
            role=RoleEnum.DOCENTE,
            nivel=nivel,
            grado=grado,
            is_active=True,
        )
        user.set_password(password)

        try:
            db.session.add(user)
            db.session.flush()
            for cid in cursos:
                db.session.add(TeacherCourse(user_id=user.id, course_id=cid))
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            result["errores"].append({
                "fila": row_num,
                "usuario": usuario,
                "motivo": "El nombre de usuario ya existe en el sistema",
            })
            continue
        except Exception:
            db.session.rollback()
            raise

        result["creados"] += 1
        result["detalle"].append({
            "fila": row_num,
            "usuario": usuario,
            "nombre": nombre,
            "password_nota": "La indicada en el Excel" if pwd_from_file else password,
            "password_es_temporal": not pwd_from_file,
            "cursos_asignados": len(cursos),
        })

    wb.close()
    return result


def generate_teachers_template_excel() -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Docentes"

    header_fill = PatternFill("solid", fgColor="0E2F77")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    example_fill = PatternFill("solid", fgColor="EEF2FF")
    example_font = Font(color="374151", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    ws.append(COLUMNAS)
    for col_num in range(1, len(COLUMNAS) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ejemplo = [
        "profe.juan",
        "Juan Pérez García",
        "MiClaveSegura1",
        "PRIMARIA",
        "3",
        "1,2",
    ]
    ws.append(ejemplo)
    for col_num in range(1, len(COLUMNAS) + 1):
        cell = ws.cell(row=2, column=col_num)
        cell.fill = example_fill
        cell.font = example_font
        cell.alignment = center
        cell.border = border

    widths = [18, 28, 18, 14, 8, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Hoja catálogo de cursos
    ws2 = wb.create_sheet("Cursos_ID")
    ws2.append(["ID", "NOMBRE", "NIVEL", "GRADO", "AREA"])
    for c in range(1, 6):
        cell = ws2.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = header_fill
        cell.border = border
    for c in Course.query.order_by(Course.nivel, Course.grado, Course.nombre).all():
        ws2.append([c.id, c.nombre, c.nivel, c.grado or "", c.area])

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = [10, 36, 12, 8, 22][col - 1]

    ws3 = wb.create_sheet("Instrucciones")
    instr = [
        ["IMPORTACIÓN DE DOCENTES"],
        [""],
        ["Columna", "Obligatorio", "Descripción"],
        ["USUARIO", "SÍ", "Nombre de login único (3-64 caracteres: letras, números, . _ -)"],
        ["NOMBRE_COMPLETO", "SÍ", "Nombre y apellidos"],
        ["PASSWORD", "NO", "Si se deja vacío, se genera una contraseña temporal (se muestra al finalizar)"],
        ["NIVEL", "NO", f"INICIAL, PRIMARIA o SECUNDARIA"],
        ["GRADO", "NO", "Si indica NIVEL, use el grado correspondiente"],
        ["IDS_CURSOS", "NO", "IDs de la hoja Cursos_ID, separados por coma (ej: 12,15,20)"],
        [""],
        ["NOTA: El rol siempre será DOCENTE."],
    ]
    for r in instr:
        ws3.append(r)
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 62

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
